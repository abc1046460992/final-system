import tkinter as tk
from tkinter import messagebox, ttk, simpledialog
import customtkinter as ctk
import threading
import os
import time
import openpyxl

from whatsapp_driver import WhatsAppDriver

import sys

# إعداد الإستايل العام لـ CustomTkinter
ctk.set_appearance_mode("Dark")  # الألوان داكنة
ctk.set_default_color_theme("green")  # اللون الأساسي أخضر مميز (مناسب للواتساب)

def fix_arabic_shortcuts(widget):
    if hasattr(widget, '_textbox'):
        native_widget = widget._textbox
    elif hasattr(widget, '_entry'):
        native_widget = widget._entry
    else:
        native_widget = widget

    def force_paste(e=None):
        try:
            native_widget.event_generate("<<Paste>>")
        except:
            pass
        return "break"

    def force_copy(e=None):
        try:
            native_widget.event_generate("<<Copy>>")
        except:
            pass
        return "break"
        
    def force_cut(e=None):
        try:
            native_widget.event_generate("<<Cut>>")
        except:
            pass
        return "break"

    def force_select_all(e=None):
        try:
            native_widget.event_generate("<<SelectAll>>")
        except:
            pass
        return "break"

    def on_keypress(event):
        # التحقق مما إذا كان زر Control مضغوطاً (event.state & 4)
        if event.state & 4:
            if event.char == 'ر': return force_paste(event)
            elif event.char == 'ؤ': return force_copy(event)
            elif event.char == 'ء': return force_cut(event)
            elif event.char == 'ش': return force_select_all(event)
            
    native_widget.bind("<KeyPress>", on_keypress, add="+")
    
    # إنشاء قائمة الزر الأيمن المنسدلة
    menu = tk.Menu(widget, tearoff=0, font=("Arial", 12))
    menu.add_command(label="لصق (Paste)", command=force_paste)
    menu.add_command(label="نسخ (Copy)", command=force_copy)
    menu.add_command(label="قص (Cut)", command=force_cut)
    menu.add_separator()
    menu.add_command(label="تحديد الكل (Select All)", command=force_select_all)

    def show_menu(event):
        try:
            # أهم نقطة: التركيز على الحقل قبل اللصق لتجنب اللصق في مكان آخر
            widget.focus()
            native_widget.focus_set()
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    native_widget.bind("<Button-3>", show_menu)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

EXCEL_FILE = "contacts.xlsx"
TEMPLATES_FILE = "templates.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(["الاسم", "رقم الجوال"])
        wb.save(EXCEL_FILE)

def init_templates():
    if not os.path.exists(TEMPLATES_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Templates"
        ws.append(["الاختصار", "نص الرسالة"])
        wb.save(TEMPLATES_FILE)

def load_contacts():
    for item in tree.get_children():
        tree.delete(item)
    
    if not os.path.exists(EXCEL_FILE):
        init_excel()
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            tree.insert("", tk.END, values=(row[0], row[1]))

def add_contact():
    name = name_entry.get().strip()
    phone = phone_entry.get().strip()
    
    if not name or not phone:
        messagebox.showwarning("تنبيه", "أدخل الاسم ورقم الجوال")
        return
        
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, phone])
    wb.save(EXCEL_FILE)
    
    name_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    phone_entry.insert(0, "+966")
    
    load_contacts()
    messagebox.showinfo("نجاح", "تمت إضافة جهة الاتصال بنجاح")

def load_templates_into_combo():
    if not os.path.exists(TEMPLATES_FILE):
        init_templates()
        template_combo.configure(values=["اضف قوالب أولاً"])
        return
    wb = openpyxl.load_workbook(TEMPLATES_FILE)
    ws = wb.active
    templates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]: # الاختصار موجود
            templates.append(str(row[0]))
    
    if templates:
        template_combo.configure(values=templates)
    else:
        template_combo.configure(values=["لا توجد قوالب محفوظة"])

def on_template_selected(choice):
    shortcut = choice
    if not shortcut or "لا توجد" in shortcut or "اضف" in shortcut: return
    wb = openpyxl.load_workbook(TEMPLATES_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == shortcut:
            msg_text.delete("1.0", tk.END)
            msg_text.insert(tk.END, row[1])
            break

def save_template():
    msg = msg_text.get("1.0", tk.END).strip()
    if not msg:
        messagebox.showwarning("تنبيه", "النص فارغ، اكتب رسالة أولاً لحفظها كقالب.")
        return
    shortcut = simpledialog.askstring("حفظ قالب", "اكتب اسماً أو اختصاراً للرسالة\n(مثال: رسالة تعميم، معايدة العيد):")
    if not shortcut:
        return
        
    shortcut = shortcut.strip()
    if not shortcut: return
    
    wb = openpyxl.load_workbook(TEMPLATES_FILE)
    ws = wb.active
    
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == shortcut:
            row[1].value = msg
            found = True
            break
            
    if not found:
        ws.append([shortcut, msg])
        
    wb.save(TEMPLATES_FILE)
    load_templates_into_combo()
    template_combo.set(shortcut)
    messagebox.showinfo("نجاح", f"تم حفظ القالب '{shortcut}' بنجاح.")

def select_all():
    tree.selection_set(tree.get_children())

def send_bulk():
    selected_items = tree.selection()
    if not selected_items:
        messagebox.showwarning("تنبيه", "الرجاء تحديد جهة اتصال واحدة على الأقل من القائمة بالماوس.")
        return
        
    msg = msg_text.get("1.0", tk.END).strip()
    if not msg:
        messagebox.showwarning("تنبيه", "أدخل نص الرسالة")
        return

    contacts_to_send = []
    for item in selected_items:
        values = tree.item(item, "values")
        contacts_to_send.append({"name": values[0], "phone": values[1]})

    def run():
        try:
            status_label.configure(text="⏳ جاري فتح واتساب...")
            send_btn.configure(state="disabled")
            progress_bar.set(0)
            percent_label.configure(text="0%")
            app.update_idletasks()

            driver = WhatsAppDriver(headless=headless_var.get())

            if not driver.load_session():
                status_label.configure(text="❌ فشل تسجيل الدخول أو إغلاق التطبيق")
                send_btn.configure(state="normal")
                return

            success_count = 0
            fail_count = 0
            total_contacts = len(contacts_to_send)

            for i, contact in enumerate(contacts_to_send, 1):
                phone = str(contact["phone"])
                name = contact["name"]
                
                status_label.configure(text=f"⏳ الإرسال إلى: {name} ({i}/{total_contacts})...")
                
                success = driver.send_message(
                    phone=phone,
                    text=msg,
                    file_path=None
                )

                if success:
                    success_count += 1
                else:
                    fail_count += 1
                    
                progress_val = (i / total_contacts)
                progress_bar.set(progress_val)
                percent_label.configure(text=f"{int(progress_val * 100)}%")
                app.update_idletasks()
                    
                time.sleep(2) # انتظار بسيط بين الرسائل

            status_label.configure(text=f"✅ اكتمل الإرسال الجماعي: {success_count} نجاح، {fail_count} فشل")
            driver.close()
            send_btn.configure(state="normal")
            
            messagebox.showinfo(
                "تقرير الإرسال", 
                f"إجمالي المحددين: {total_contacts}\n"
                f"✅ تم الإرسال بنجاح: {success_count}\n"
                f"❌ المفقودين/فشل الاتصال: {fail_count}"
            )

        except Exception as e:
            messagebox.showerror("خطأ", str(e))
            status_label.configure(text="❌ خطأ أثناء الإرسال")
            send_btn.configure(state="normal")

    threading.Thread(target=run).start()

def send_single():
    phone = single_phone_entry.get().strip()
    msg = single_msg_text.get("1.0", tk.END).strip()
    
    if not phone or not msg:
        messagebox.showwarning("تنبيه", "الرجاء إدخال رقم الجوال ونص الرسالة")
        return

    def run_single():
        try:
            single_status_label.configure(text="⏳ جاري فتح واتساب للرسالة السريعة...")
            single_send_btn.configure(state="disabled")

            driver = WhatsAppDriver(headless=headless_var.get())

            if not driver.load_session():
                single_status_label.configure(text="❌ فشل تسجيل الدخول")
                single_send_btn.configure(state="normal")
                return

            single_status_label.configure(text=f"⏳ جاري إرسال الرسالة إلى {phone}...")
            
            success = driver.send_message(
                phone=phone,
                text=msg,
                file_path=None
            )

            if success:
                single_status_label.configure(text="✅ تم الإرسال بنجاح لـ " + phone)
                time.sleep(2)  # انتظار بسيط بعد الإرسال
                messagebox.showinfo("نجاح", "تم إرسال الرسالة السريعة بنجاح! 🚀")
                single_phone_entry.delete(0, tk.END)
                single_phone_entry.insert(0, "+966")
                single_msg_text.delete("1.0", tk.END)
            else:
                single_status_label.configure(text="❌ فشل الإرسال، الرقم غير مدعوم ربما.")
                messagebox.showerror("خطأ", "فشل الإرسال، الرجاء التأكد من الرقم والاتصال.")
                
            driver.close()
            single_send_btn.configure(state="normal")

        except Exception as e:
            messagebox.showerror("خطأ داخلي", str(e))
            single_status_label.configure(text="❌ خطأ أثناء الإرسال")
            single_send_btn.configure(state="normal")

    threading.Thread(target=run_single).start()

# =========================================================
# ==================== واجهة المستخدم ====================
# =========================================================

app = ctk.CTk()
app.title("واتساب بوت 🤖 - الإصدار الاحترافي الكامل 🚀")
icon_path = resource_path("whatsapp.ico")
if os.path.exists(icon_path):
    app.iconbitmap(icon_path)
app.geometry("650x850")
app.configure(fg_color="#0f172a")

# إعداد ملفات قاعدة البيانات
init_excel()
init_templates()

# متغير التخفي المشترك (يعمل في كل التبويبات)
headless_var = ctk.BooleanVar(value=False)

# ----------------- إعداد التبويبات الفخمة -----------------
tabview = ctk.CTkTabview(app, corner_radius=15, fg_color="#1e293b", segmented_button_fg_color="#0f172a", segmented_button_selected_color="#2b82bb")
tabview.pack(fill="both", expand=True, padx=20, pady=20)

tab_bulk = tabview.add("👥 مجموعة الأصدقاء")
tab_single = tabview.add("📩 رسالة البرق (سريعة)")
tabview.set("👥 مجموعة الأصدقاء")

# ----------------- تنسيق الجدول (Treeview) الليلي -----------------
# نظراً لأن CustomTkinter لا يوفر جدولاً جاهزاً حتى الآن، نقوم بتزيين جدول tk ليتلاءم مع النمط الداكن الخرافي!
style = ttk.Style()
style.theme_use("default")
style.configure("Treeview", 
                background="#1e293b",
                foreground="white",
                rowheight=35,
                fieldbackground="#1e293b",
                borderwidth=0,
                font=("Arial", 12))
style.map('Treeview', background=[('selected', '#2fa572')])

style.configure("Treeview.Heading", 
                background="#0f172a",
                foreground="white",
                relief="flat",
                font=("Arial", 13, "bold"),
                padding=5)
style.map("Treeview.Heading", background=[('active', '#444444')])

# ==================== التبويب الأول (الأصدقاء) ====================

# 1- إطار إضافة جهات الاتصال بطريقة أنيقة
add_frame = ctk.CTkFrame(tab_bulk, corner_radius=12, fg_color="#1e293b")
add_frame.pack(fill="x", padx=10, pady=10)

add_inner = ctk.CTkFrame(add_frame, fg_color="transparent")
add_inner.pack(pady=10, padx=10)

ctk.CTkLabel(add_inner, text="الاسم:", font=("Arial", 14, "bold"), text_color="#A9A9A9").grid(row=0, column=0, pady=10, padx=10, sticky="e")
name_entry = ctk.CTkEntry(add_inner, width=200, font=("Arial", 13), border_width=2, corner_radius=8)
name_entry.grid(row=0, column=1, pady=10, padx=10)
fix_arabic_shortcuts(name_entry)

ctk.CTkLabel(add_inner, text="رقم الجوال:", font=("Arial", 14, "bold"), text_color="#A9A9A9").grid(row=1, column=0, pady=10, padx=10, sticky="e")
phone_entry = ctk.CTkEntry(add_inner, width=200, font=("Arial", 13), border_width=2, corner_radius=8)
phone_entry.insert(0, "+966")
phone_entry.grid(row=1, column=1, pady=10, padx=10)
fix_arabic_shortcuts(phone_entry)

add_btn = ctk.CTkButton(add_inner, text="➕ إضافة جهة", command=add_contact, corner_radius=8, font=("Arial", 14, "bold"), fg_color="#2b82bb", hover_color="#1c6494")
add_btn.grid(row=0, column=2, rowspan=2, padx=20, ipady=8)

# 2- إطار القائمة المخصصة
list_frame = ctk.CTkFrame(tab_bulk, corner_radius=12, fg_color="#1e293b")
list_frame.pack(fill="both", expand=True, padx=10, pady=5)

ctk.CTkLabel(list_frame, text="قائمة الأصدقاء المحفوظين (انقر لتحديد المستلمين أو استخدم زر التحديد)", font=("Arial", 11), text_color="gray").pack(pady=(10,0))

tree_scroll = ctk.CTkScrollbar(list_frame)
tree_scroll.pack(side="left", fill="y", pady=10, padx=(10,0))

tree = ttk.Treeview(list_frame, columns=("Name", "Phone"), show="headings", selectmode="extended", yscrollcommand=tree_scroll.set)
tree.heading("Name", text="🎭 الاسم")
tree.heading("Phone", text="📞 رقم الجوال")
tree.column("Name", width=250, anchor="center")
tree.column("Phone", width=250, anchor="center")
tree.pack(fill="both", expand=True, padx=10, pady=10)
tree_scroll.configure(command=tree.yview)

btn_frame = ctk.CTkFrame(list_frame, fg_color="transparent")
btn_frame.pack(fill="x", pady=5, padx=10)
select_all_btn = ctk.CTkButton(btn_frame, text="✅ تحديد الكل", command=select_all, corner_radius=8, font=("Arial", 12, "bold"), fg_color="#e67e22", hover_color="#d35400")
select_all_btn.pack(side="right")

# 3- إطار الرسالة وتحديد القوالب
msg_frame = ctk.CTkFrame(tab_bulk, corner_radius=12, fg_color="#1e293b")
msg_frame.pack(fill="x", padx=10, pady=10)

template_frame = ctk.CTkFrame(msg_frame, fg_color="transparent")
template_frame.pack(fill="x", pady=10, padx=10)

save_tpl_btn = ctk.CTkButton(template_frame, text="📥 حفظ النص كقالب", command=save_template, corner_radius=8, width=120, font=("Arial", 12, "bold"), fg_color="#8e44ad", hover_color="#732d91")
save_tpl_btn.pack(side="left")

template_combo = ctk.CTkOptionMenu(template_frame, width=200, command=on_template_selected, font=("Arial", 13, "bold"), dropdown_font=("Arial", 12), corner_radius=8)
template_combo.pack(side="right")
template_combo.set("اختر قالب محفوظ...")

ctk.CTkLabel(template_frame, text="قوالب الرسائل الجاهزة:", font=("Arial", 14, "bold")).pack(side="right", padx=10)

msg_text = ctk.CTkTextbox(msg_frame, height=100, font=("Arial", 15), corner_radius=10, border_width=1, border_color="#333333")
msg_text.pack(fill="both", padx=10, pady=5)
fix_arabic_shortcuts(msg_text)

# 4- خيارات وأزرار التحكم
options_frame = ctk.CTkFrame(tab_bulk, fg_color="transparent")
options_frame.pack(fill="x", padx=10, pady=(5,0))

headless_switch = ctk.CTkSwitch(options_frame, text="متصفح صامت / بالخلفية 🕶️", variable=headless_var, font=("Arial", 13, "bold"), onvalue=True, offvalue=False)
headless_switch.pack(side="right")

send_btn = ctk.CTkButton(tab_bulk, text="🚀 إرسال الرسالة للمحددين", command=send_bulk, corner_radius=20, font=("Arial", 18, "bold"), height=55)
send_btn.pack(pady=15, padx=60, fill="x")

# شريط الإنجاز المتطور
progress_frame = ctk.CTkFrame(tab_bulk, fg_color="transparent", height=30)
progress_frame.pack(fill="x", padx=40, pady=5)

progress_bar = ctk.CTkProgressBar(progress_frame, height=22, corner_radius=10, progress_color="#2ecc71")
progress_bar.set(0)
progress_bar.place(relx=0.5, rely=0.5, anchor="center", relwidth=1)

percent_label = ctk.CTkLabel(progress_frame, text="0%", font=("Arial", 11, "bold"), text_color="#1E1E1E")
percent_label.place(relx=0.5, rely=0.5, anchor="center")

status_label = ctk.CTkLabel(tab_bulk, text="مُستعد للإرسال للمجموعة", text_color="gray", font=("Arial", 12))
status_label.pack()

# ==================== التبويب الثاني (تجربة لمرة واحدة) ====================

single_inner = ctk.CTkFrame(tab_single, corner_radius=15, fg_color="#1e293b")
single_inner.pack(fill="both", expand=True, padx=20, pady=20)

ctk.CTkLabel(single_inner, text="المرسال السريع ✉️", font=("Arial", 24, "bold"), text_color="#2fa572").pack(pady=(30, 10))
ctk.CTkLabel(single_inner, text="أرسل رسالة سريعة دون الحاجة لحفظ رقم المستلم\nأو تكرار خطوات التعبئة. مثالي للأشخاص المؤقتين.", text_color="gray", font=("Arial", 14)).pack(pady=(0,30))

ctk.CTkLabel(single_inner, text="رقم الجوال:", font=("Arial", 16, "bold")).pack(anchor="e", pady=(10,5), padx=40)
single_phone_entry = ctk.CTkEntry(single_inner, width=320, height=50, font=("Arial", 20, "bold"), justify="center", corner_radius=12)
single_phone_entry.insert(0, "+966")
single_phone_entry.pack(pady=5)
fix_arabic_shortcuts(single_phone_entry)

ctk.CTkLabel(single_inner, text="نص الرسالة المطلوبة:", font=("Arial", 16, "bold")).pack(anchor="e", pady=(25, 5), padx=40)
single_msg_text = ctk.CTkTextbox(single_inner, width=500, height=200, font=("Arial", 16), corner_radius=12, border_width=1, border_color="#333333")
single_msg_text.pack(pady=5)
fix_arabic_shortcuts(single_msg_text)

s_options_frame = ctk.CTkFrame(single_inner, fg_color="transparent")
s_options_frame.pack(fill="x", pady=20, padx=50)
single_headless_switch = ctk.CTkSwitch(s_options_frame, text="تشغيل صامت 🕶️", variable=headless_var, font=("Arial", 13, "bold"), onvalue=True, offvalue=False)
single_headless_switch.pack(side="right")

single_send_btn = ctk.CTkButton(single_inner, text="🚀 إطلاق الرسالة فوراً!", command=send_single, corner_radius=18, font=("Arial", 20, "bold"), height=65, fg_color="#c0392b", hover_color="#962d22")
single_send_btn.pack(pady=20, padx=60, fill="x")

single_status_label = ctk.CTkLabel(single_inner, text="مُستعد للإرسال الفردي", text_color="gray", font=("Arial", 12))
single_status_label.pack(pady=10)

# ==================== التشغيل النهائي ====================
info_frame = ctk.CTkFrame(app, fg_color="transparent")
info_frame.pack(side="bottom", fill="x", pady=(0, 10))
ctk.CTkLabel(info_frame, text="💡 تلميح ودّي: لضمان نجاح الإرسال، درعك الأول هو التأكد من أن حساب الواتساب الخاص بك مسجل الدخول،\nوأن جهازك متصل بالإنترنت. 🛡️", text_color="#7FA1C3", font=("Arial", 12, "bold")).pack()

load_contacts()
load_templates_into_combo()
app.mainloop()